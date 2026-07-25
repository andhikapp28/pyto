"""
Bab 15: Asisten Kantor Kilat (Excel & PDF Sederhana) 📊
=========================================================
Punya file Excel yang tiap bulan harus dijumlahkan dan disortir manual?
Sekarang giliran Python yang kerja: baca file Excel, jumlahkan angka,
pilih baris yang penting, lalu tulis ulang jadi file Excel baru.

Catatan: berkas ini butuh paket openpyxl (untuk Excel) dan pypdf
(untuk bagian bonus PDF). Install dulu lewat terminal:
    pip install openpyxl pypdf

Supaya berkas ini bisa langsung dijalankan dari nol tanpa perlu file
Excel siapan sendiri, bagian pertama di bawah akan MEMBUAT contoh file
"data_penjualan.xlsx" dulu (data toko fiktif) - baru setelah itu
kodenya membaca file tersebut persis seperti membaca file Excel
sungguhan. Kalau kamu punya file Excel sendiri, tinggal ganti nama
filenya di baris openpyxl.load_workbook(...) di bawah.

Soal PDF (bagian "Tahu Lebih" di akhir): taruh file PDF apa saja
bernama "contoh.pdf" di folder yang sama dengan berkas ini kalau mau
mencobanya - kalau filenya tidak ada, bagian itu otomatis dilewati
dengan pesan ramah (bukan bikin program berhenti/error).
"""

import openpyxl

# Siapkan dulu contoh file Excel-nya (data toko fiktif) --------------
# Di dunia nyata, langkah ini tidak perlu - kamu tinggal punya file
# Excel-mu sendiri. Ini cuma supaya berkas ini 100% bisa dijalankan
# dari nol tanpa aset dari luar.
buku_contoh = openpyxl.Workbook()
sheet_contoh = buku_contoh.active
sheet_contoh.append(["Nama Barang", "Jumlah Terjual", "Harga Satuan"])
sheet_contoh.append(["Kopi Sachet", 120, 2500])
sheet_contoh.append(["Teh Celup", 80, 1500])
sheet_contoh.append(["Gula Pasir 1kg", 40, 13000])
sheet_contoh.append(["Mie Instan", 200, 3000])
sheet_contoh.append(["Sabun Mandi", 30, 4000])
sheet_contoh.append(["Pasta Gigi", 15, 8000])
sheet_contoh.append(["Permen Karet", 50, 1000])
sheet_contoh.append(["Tisu Basah", 10, 12000])
sheet_contoh.append(["Air Mineral 600ml", 100, 3500])
sheet_contoh.append(["Korek Api", 20, 500])
buku_contoh.save("data_penjualan.xlsx")
print("Contoh file data_penjualan.xlsx sudah dibuat.")


# Baca isi Excel jadi list berisi dictionary ---------------------------
buku = openpyxl.load_workbook("data_penjualan.xlsx")
sheet = buku.active

data_penjualan = []
for baris in sheet.iter_rows(min_row=2, values_only=True):
    nama, jumlah, harga = baris
    data_penjualan.append({"nama": nama, "jumlah": jumlah, "harga": harga})

print(data_penjualan[0])


# ⚠️ Awas - baris pertama itu biasanya judul kolom, bukan data!
# Kalau kita lupa min_row=2, baris judul ("Nama Barang", dst.) ikut
# terbaca seolah data sungguhan, dan program akan berhenti begitu
# mencoba mengalikan dua teks:
#
#   for baris in sheet.iter_rows(values_only=True):
#       nama, jumlah, harga = baris
#       total = jumlah * harga   # 💥 error di baris pertama:
#       # TypeError: can't multiply sequence by non-int of type 'str'
#
# Sengaja dikomentari supaya berkas ini tetap bisa dijalankan sampai
# selesai. Coba hapus tanda pagar di atas sendiri untuk melihat errornya.


# Olah datanya: jumlahkan & pilih yang penting -------------------------
barang_laris = []
for barang in data_penjualan:
    total = barang["jumlah"] * barang["harga"]
    barang["total"] = total
    if total >= 100000:
        barang_laris.append(barang)

print(f"Ada {len(barang_laris)} barang laris dari {len(data_penjualan)} barang.")


# Tulis hasilnya jadi file Excel baru -----------------------------------
buku_baru = openpyxl.Workbook()
sheet_baru = buku_baru.active
sheet_baru.append(["Nama Barang", "Jumlah Terjual", "Harga Satuan", "Total"])

for barang in barang_laris:
    sheet_baru.append([barang["nama"], barang["jumlah"], barang["harga"], barang["total"]])

buku_baru.save("hasil_penjualan_laris.xlsx")
print("Hasil olahan disimpan sebagai hasil_penjualan_laris.xlsx")


# 🎮 Main Yuk! - ganti batas "laris"
# Coba angka batas yang lebih longgar atau lebih ketat.
batas_laris = 50000   # <- ganti angka ini sesuka hatimu

barang_laris_main_yuk = []
for barang in data_penjualan:
    total = barang["jumlah"] * barang["harga"]
    if total >= batas_laris:
        barang_laris_main_yuk.append(barang["nama"])

print(f"Dengan batas Rp{batas_laris:,}, barang laris: {barang_laris_main_yuk}")


# 🔍 Tahu Lebih - ekstrak teks dari file PDF
# Taruh file PDF apa saja bernama "contoh.pdf" di folder ini kalau mau
# mencoba bagian ini. Kalau filenya belum ada, kita cuma tampilkan
# pesan ramah dan lanjut - bukan bikin program berhenti.
try:
    import pypdf

    pembaca = pypdf.PdfReader("contoh.pdf")

    teks_semua = ""
    for halaman in pembaca.pages:
        teks_semua += halaman.extract_text()

    print(teks_semua[:500])
except FileNotFoundError:
    print("Lewati contoh PDF: taruh contoh.pdf di folder ini untuk mencoba.")
